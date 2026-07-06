from django.contrib import admin
from django.urls import path
from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from django import forms
from django.core.exceptions import ValidationError
from django.db.models import Sum
import openpyxl

from .models import (
    Course, CourseAssessment, Student, Group, Rubric, RubricLevel, 
    RubricItem, Evaluation, GroupItemScore, IndividualItemScore
)

import json

class ImportStudentsForm(forms.Form):
    excel_file = forms.FileField(label="Excel File (.xlsx)")

class ImportRubricForm(forms.Form):
    json_file = forms.FileField(label="JSON File (.json)")

class CourseAssessmentInlineFormSet(forms.models.BaseInlineFormSet):
    def clean(self):
        super().clean()
        if any(self.errors):
            return
        
        pres_weight = 0
        exam_weight = 0
        has_pres = False
        has_exam = False

        for form in self.forms:
            if self.can_delete and self._should_delete_form(form):
                continue
            weight = form.cleaned_data.get('weight_percentage')
            atype = form.cleaned_data.get('assessment_type')
            if weight is not None and atype:
                if atype == 'PRESENTATION':
                    pres_weight += weight
                    has_pres = True
                elif atype == 'EXAM':
                    exam_weight += weight
                    has_exam = True
                
        if has_pres and abs(pres_weight - 100) > 0.01:
            raise ValidationError(f"The sum of PRESENTATION weight percentages must be exactly 100. Current sum: {pres_weight}%")
        
        if has_exam and abs(exam_weight - 100) > 0.01:
            raise ValidationError(f"The sum of EXAM weight percentages must be exactly 100. Current sum: {exam_weight}%")

class CourseAssessmentInline(admin.TabularInline):
    model = CourseAssessment
    extra = 1
    formset = CourseAssessmentInlineFormSet

class CourseAdminForm(forms.ModelForm):
    class Meta:
        model = Course
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        p_weight = cleaned_data.get('presentation_weight')
        e_weight = cleaned_data.get('exam_weight')
        if p_weight is not None and e_weight is not None:
            if abs(p_weight + e_weight - 100) > 0.01:
                raise ValidationError("Presentation weight and Exam weight must sum to exactly 100.")
        return cleaned_data

@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    form = CourseAdminForm
    list_display = ('code', 'name', 'section', 'period')
    search_fields = ('code', 'name')
    inlines = [CourseAssessmentInline]

    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:object_id>/import-students/',
                self.admin_site.admin_view(self.import_students_view),
                name='course-import-students',
            ),
            path(
                '<path:object_id>/import-rubric/',
                self.admin_site.admin_view(self.import_rubric_view),
                name='course-import-rubric',
            ),
        ]
        return custom_urls + urls

    def import_rubric_view(self, request, object_id):
        course = get_object_or_404(Course, pk=object_id)
        if request.method == "POST":
            form = ImportRubricForm(request.POST, request.FILES)
            if form.is_valid():
                json_file = form.cleaned_data['json_file']
                try:
                    data = json.load(json_file)
                    
                    # 1. Create Rubric
                    metadata = data.get('rubric_metadata', {})
                    rubric = Rubric.objects.create(
                        course=course,
                        name=metadata.get('name', 'Imported Rubric'),
                        description=metadata.get('description', '')
                    )
                    
                    # 2. Create Levels
                    levels_data = data.get('levels', [])
                    level_objs = {}
                    for lvl in levels_data:
                        cat_name = lvl.get('category_name')
                        level_obj = RubricLevel.objects.create(
                            rubric=rubric,
                            category_name=cat_name,
                            score_percentage=lvl.get('score_percentage', 0)
                        )
                        level_objs[cat_name] = level_obj
                        
                    # 3. Create Items and Cells
                    from .models import RubricCell
                    items_data = data.get('items', [])
                    for item_d in items_data:
                        item_obj = RubricItem.objects.create(
                            rubric=rubric,
                            order=item_d.get('order', ''),
                            item_type=item_d.get('item_type', 'GROUP'),
                            description=item_d.get('description', ''),
                            weight=item_d.get('weight', 0.1)
                        )
                        
                        cells_data = item_d.get('cells', {})
                        for cat_name, desc in cells_data.items():
                            if cat_name in level_objs:
                                RubricCell.objects.create(
                                    item=item_obj,
                                    level=level_objs[cat_name],
                                    description=desc
                                )
                                
                    messages.success(request, f"Rubric '{rubric.name}' imported successfully into course {course.code}.")
                    return HttpResponseRedirect("..")
                    
                except json.JSONDecodeError:
                    messages.error(request, "Invalid JSON file.")
                    return HttpResponseRedirect(".")
                except Exception as e:
                    messages.error(request, f"Error processing JSON file: {e}")
                    return HttpResponseRedirect(".")
        else:
            form = ImportRubricForm()
            
        context = self.admin_site.each_context(request)
        context['opts'] = self.model._meta
        context['form'] = form
        context['course'] = course
        context['title'] = f"Import Rubric for {course}"
        return render(request, "admin/evaluations/course/import_rubric.html", context)

    def import_students_view(self, request, object_id):
        course = get_object_or_404(Course, pk=object_id)
        if request.method == "POST":
            form = ImportStudentsForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data['excel_file']
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    sheet = wb.active
                    
                    header_row = None
                    rut_col, pat_col, mat_col, nom_col = None, None, None, None
                    
                    # Find headers dynamically
                    for row in sheet.iter_rows(min_row=1, max_row=50):
                        for cell in row:
                            if cell.value and str(cell.value).strip().lower() == 'rut alumno':
                                header_row = cell.row
                                break
                        if header_row:
                            break
                            
                    if not header_row:
                        messages.error(request, "Could not find 'Rut Alumno' header in the first 50 rows.")
                        return HttpResponseRedirect("..")
                        
                    # Map columns
                    for cell in sheet[header_row]:
                        val = str(cell.value).strip().lower() if cell.value else ''
                        if val == 'rut alumno': rut_col = cell.column
                        elif val == 'apellido paterno': pat_col = cell.column
                        elif val == 'apellido materno': mat_col = cell.column
                        elif val == 'nombre': nom_col = cell.column
                        
                    if not all([rut_col, pat_col, nom_col]):
                        messages.error(request, "Missing required columns (Rut Alumno, Apellido Paterno, Nombre).")
                        return HttpResponseRedirect("..")
                        
                    created_count = 0
                    enrolled_count = 0
                    
                    for row in sheet.iter_rows(min_row=header_row + 1):
                        rut = row[rut_col-1].value
                        if not rut:
                            continue # Skip empty rows
                            
                        rut = str(rut).strip()
                        pat = str(row[pat_col-1].value or '').strip()
                        mat = str(row[mat_col-1].value or '').strip() if mat_col else ''
                        nom = str(row[nom_col-1].value or '').strip()
                        
                        # Handle student creation/retrieval
                        student, created = Student.objects.get_or_create(
                            identification=rut,
                            defaults={
                                'first_names': nom,
                                'last_name_1': pat,
                                'last_name_2': mat
                            }
                        )
                        if created:
                            created_count += 1
                            
                        # Handle enrollment safely
                        if not student.courses.filter(id=course.id).exists():
                            student.courses.add(course)
                            enrolled_count += 1
                            
                    messages.success(
                        request, 
                        f"Import successful! Created {created_count} new students. Enrolled {enrolled_count} students in {course.code}."
                    )
                    return HttpResponseRedirect("..")
                    
                except Exception as e:
                    messages.error(request, f"Error processing Excel file: {e}")
                    return HttpResponseRedirect(".")
        else:
            form = ImportStudentsForm()
            
        context = self.admin_site.each_context(request)
        context['opts'] = self.model._meta
        context['form'] = form
        context['course'] = course
        context['title'] = f"Import Students for {course}"
        return render(request, "admin/evaluations/course/import_students.html", context)

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = ('identification', 'last_name_1', 'last_name_2', 'first_names')
    search_fields = ('identification', 'last_name_1', 'first_names')
    filter_horizontal = ('courses',)

@admin.register(Group)
class GroupAdmin(admin.ModelAdmin):
    list_display = ('name', 'course')
    list_filter = ('course',)
    filter_horizontal = ('members',)

class RubricLevelInline(admin.TabularInline):
    model = RubricLevel
    extra = 3

class RubricItemInline(admin.TabularInline):
    model = RubricItem
    extra = 3

@admin.register(Rubric)
class RubricAdmin(admin.ModelAdmin):
    list_display = ('name', 'course')
    list_filter = ('course',)
    inlines = [RubricLevelInline, RubricItemInline]
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:object_id>/matrix/',
                self.admin_site.admin_view(self.matrix_view),
                name='rubric-matrix',
            ),
        ]
        return custom_urls + urls

    def matrix_view(self, request, object_id):
        rubric = get_object_or_404(Rubric, pk=object_id)
        levels = rubric.levels.all()
        items = rubric.items.all()
        
        from .models import RubricCell
        
        if request.method == "POST":
            for item in items:
                for level in levels:
                    cell_key = f"cell_{item.id}_{level.id}"
                    desc = request.POST.get(cell_key, "").strip()
                    if desc:
                        RubricCell.objects.update_or_create(
                            item=item, level=level,
                            defaults={'description': desc}
                        )
                    else:
                        RubricCell.objects.filter(item=item, level=level).delete()
            messages.success(request, "Rubric matrix updated successfully.")
            return HttpResponseRedirect("..")
            
        cells = RubricCell.objects.filter(item__rubric=rubric)
        
        # Build a list of rows for the template to avoid complex dict lookups in Django templates
        matrix_rows = []
        for item in items:
            row_cells = []
            for level in levels:
                cell = cells.filter(item=item, level=level).first()
                row_cells.append({
                    'level': level,
                    'key': f"cell_{item.id}_{level.id}",
                    'description': cell.description if cell else ""
                })
            matrix_rows.append({
                'item': item,
                'cells': row_cells
            })

        context = self.admin_site.each_context(request)
        context.update({
            'rubric': rubric,
            'levels': levels,
            'matrix_rows': matrix_rows,
            'title': f"Edit Matrix for {rubric.name}",
            'opts': self.model._meta,
        })
        return render(request, "admin/evaluations/rubric/matrix_editor.html", context)

class GroupItemScoreInline(admin.TabularInline):
    model = GroupItemScore
    extra = 0

class IndividualItemScoreInline(admin.TabularInline):
    model = IndividualItemScore
    extra = 0

from django.utils.html import format_html

@admin.register(Evaluation)
class EvaluationAdmin(admin.ModelAdmin):
    list_display = ('assessment', 'group', 'date', 'grade_link')
    list_filter = ('assessment', 'group')
    filter_horizontal = ('evaluated_students',)
    inlines = [GroupItemScoreInline, IndividualItemScoreInline]
    
    def grade_link(self, obj):
        return format_html(
            '<a class="button" href="/evaluation/{}/grade/">Grade Interactively</a>',
            obj.id
        )
    grade_link.short_description = 'Interactive Grading'
    grade_link.allow_tags = True

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        # Snapshot the students when the evaluation is created
        if not change and not form.instance.evaluated_students.exists():
            form.instance.evaluated_students.set(form.instance.group.members.all())


